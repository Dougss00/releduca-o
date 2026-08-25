from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


def build_summary_workbook(data: dict) -> bytes:
    """Gera o relatório horizontal a partir do resultado já consolidado."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Quantitativos"

    blue = "0868C9"
    blue_dark = "064B91"
    blue_soft = "EAF3FC"
    yellow = "F5B400"
    white = "FFFFFF"
    thin = Side(style="thin", color="C9D6E2")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=1)
    sheet["A1"] = "Estabelecimento"
    column = 2
    for grade in data["series"]:
        sheet.merge_cells(start_row=1, start_column=column, end_row=1, end_column=column + 2)
        sheet.cell(1, column, grade)
        for offset, label in enumerate(("F", "M", "Total")):
            sheet.cell(2, column + offset, label)
        column += 3
    sheet.merge_cells(start_row=1, start_column=column, end_row=1, end_column=column + 2)
    sheet.cell(1, column, "Totais")
    for offset, label in enumerate(("F", "M", "Geral")):
        sheet.cell(2, column + offset, label)

    sheet["A2"] = ""
    last_column = column + 2
    for row in sheet.iter_rows(min_row=1, max_row=2, min_col=1, max_col=last_column):
        for cell in row:
            cell.fill = PatternFill("solid", fgColor=blue_dark if cell.row == 1 else blue)
            cell.font = Font(color=white, bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

    for row_number, item in enumerate(data["estabelecimentos"], start=3):
        sheet.cell(row_number, 1, item["estabelecimento"])
        column = 2
        for grade in data["series"]:
            values = item["series"][grade]
            for value in (values["F"], values["M"], values["total"]):
                sheet.cell(row_number, column, value)
                column += 1
        for value in (item["total_f"], item["total_m"], item["total_geral"]):
            sheet.cell(row_number, column, value)
            column += 1
        for cell in sheet[row_number]:
            cell.border = border
            cell.alignment = Alignment(horizontal="left" if cell.column == 1 else "center")
            if cell.column >= last_column - 2 or (cell.column > 1 and (cell.column - 1) % 3 == 0):
                cell.fill = PatternFill("solid", fgColor=blue_soft)
                cell.font = Font(bold=True)

    total_row = sheet.max_row + 1
    sheet.cell(total_row, 1, "TOTAL GERAL")
    for column in range(2, last_column + 1):
        letter = get_column_letter(column)
        sheet.cell(total_row, column, f"=SUM({letter}3:{letter}{total_row - 1})")
    for cell in sheet[total_row]:
        cell.fill = PatternFill("solid", fgColor=yellow)
        cell.font = Font(color="1C2C3A", bold=True)
        cell.border = border
        cell.alignment = Alignment(horizontal="center")

    sheet.column_dimensions["A"].width = 58
    for column in range(2, last_column + 1):
        sheet.column_dimensions[get_column_letter(column)].width = 11
    sheet.row_dimensions[1].height = 28
    sheet.row_dimensions[2].height = 23
    sheet.freeze_panes = "B3"
    sheet.auto_filter.ref = f"A2:{get_column_letter(last_column)}{total_row - 1}"
    sheet.sheet_view.showGridLines = False
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.print_title_rows = "1:2"

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()
